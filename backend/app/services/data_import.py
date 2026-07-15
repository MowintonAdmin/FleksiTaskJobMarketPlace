"""
Historical Data Import Service

Parses Excel workbooks, validates, normalizes, detects duplicates,
and imports worker profiles and task sessions into the database.

Architecture:
  - ImportService orchestrates the full pipeline
  - RowParser handles Excel → normalized row conversion
  - Validator checks data quality
  - DuplicateDetector queries existing records
  - Importer executes the actual DB writes

The preview step (parse + validate + detect) makes NO database changes.
The confirm step wraps everything in a single transaction.
"""
import io
import uuid
import logging
import time
from datetime import datetime, timezone, date
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_

from app.models.user import User
from app.models.task import Task, TaskStatus
from app.models.application import Application
from app.models.task_session import TaskSession, SessionStatus
from app.models.import_log import ImportLog
from app.models.enums import DataSource
from app.schemas.data_import import (
    ImportRow,
    ImportPreviewResponse,
    ImportPreviewRow,
    ImportConfirmResponse,
    WorkerPreview,
)
from app.services.import_helpers import generate_import_email, parse_import_participant_id, parse_import_phone
from app.config import get_settings

logger = logging.getLogger(__name__)
settings = get_settings()

# ── Column mapping for the Excel workbook ──────────────────────────────────
COL_PARTICIPANT_ID = 1
COL_NRIC_PASSPORT = 2
COL_DATE = 3
COL_NO = 4
COL_PROJECT_ID = 5
COL_PROJECT_SESSION = 6
COL_BODY_HEIGHT = 7
COL_NATURE_OF_WORK = 8
COL_ACTIVITY_TASK = 9
COL_ENVIRONMENT = 10
COL_DEVICE_ID = 11
COL_QC_DURATION = 13       # Column M
COL_EXPECTED_DURATION = 14  # Column N
COL_TOTAL_DURATION = 15     # Column O (M+N, formula)
COL_EXPECTED_TARGET = 16    # Column P
COL_VARIANCE = 17           # Column Q
COL_RUNNING_TOTAL = 18      # Column R
COL_AVG_PERIOD = 19         # Column S
COL_SHORT_NAME = 20         # Column T
COL_FULL_NAME = 21          # Column U
COL_HOURS_WORKED = 22       # Column V (O/60, formula)
COL_TOTAL_PAYOUT = 23       # Column W
COL_PHONE = 24              # Column X

# Fixed UUID namespace for placeholder records used by imported sessions
# This deterministic ID will be created on first import and reused thereafter
PLACEHOLDER_TASK_ID = uuid.UUID("00000000-0000-0000-0000-000000000001")
PLACEHOLDER_APP_ID = uuid.UUID("00000000-0000-0000-0000-000000000002")

# Map column numbers to human-readable names for error messages
COLUMN_NAMES = {
    COL_PARTICIPANT_ID: "Participant ID",
    COL_NRIC_PASSPORT: "NRIC/Passport",
    COL_DATE: "Date",
    COL_PROJECT_ID: "Project ID",
    COL_PROJECT_SESSION: "Project Session",
    COL_FULL_NAME: "Full Name",
    COL_TOTAL_DURATION: "Total Duration",
    COL_TOTAL_PAYOUT: "Total Payout",
    COL_PHONE: "Phone",
    COL_DEVICE_ID: "Device ID",
}


def _safe_float(value) -> Optional[float]:
    """Convert a value to float, returning None if invalid."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        cleaned = str(value).strip().replace(",", "")
        if cleaned in ("", "#N/A", "N/A", "-", "0"):
            return None
        f = float(cleaned)
        if f > 10000:  # Looks like an Excel date serial, not minutes
            return None
        return f
    except (ValueError, TypeError):
        return None


def _safe_date(value) -> Optional[date]:
    """Convert a value to a date object."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        from datetime import timedelta
        try:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=float(value))).date()
        except (ValueError, OverflowError):
            return None
    try:
        cleaned = str(value).strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
    except (ValueError, TypeError):
        pass
    return None


def _normalize_name(raw: Optional[str]) -> Optional[str]:
    """Clean and normalize a person's name.
    Rejects numeric values, short codes, and single characters that are 
    actually hours-worked values rather than names.
    """
    if not raw:
        return None
    cleaned = str(raw).strip()
    if cleaned in ("", "#N/A", "N/A", "-"):
        return None
    import re
    cleaned = re.sub(r'\s*\(\d+\)\s*$', '', cleaned).strip()
    # Reject pure numeric values (these are actually hours/payout, not names)
    # Also reject single-digit names that look like they could be data column values
    if cleaned.isdigit() or (len(cleaned) <= 2 and cleaned.replace('.','').isdigit()):
        return None
    return cleaned if cleaned else None


def _normalize_nric(raw: Optional[str]) -> Optional[str]:
    """Clean an NRIC/Passport number."""
    if not raw:
        return None
    cleaned = str(raw).strip()
    if cleaned in ("", "#N/A", "N/A", "-", "0"):
        return None
    return cleaned


class ExcelRowParser:
    """Parses a single Excel worksheet row into a normalized ImportRow."""

    @staticmethod
    def parse_row(row_values: tuple, row_number: int) -> ImportRow:
        values = list(row_values) if row_values else []

        def v(idx: int):
            return values[idx - 1] if len(values) >= idx else None

        participant_id = parse_import_participant_id(v(COL_PARTICIPANT_ID))
        nric = _normalize_nric(v(COL_NRIC_PASSPORT))
        full_name = _normalize_name(v(COL_FULL_NAME)) or _normalize_name(v(COL_SHORT_NAME))
        short_name = _normalize_name(v(COL_SHORT_NAME))
        phone = parse_import_phone(v(COL_PHONE))
        raw_date = v(COL_DATE)
        parsed_date = _safe_date(raw_date)
        project_id = parse_import_participant_id(v(COL_PROJECT_ID))
        project_session = parse_import_participant_id(v(COL_PROJECT_SESSION))
        body_height = _safe_float(v(COL_BODY_HEIGHT))
        nature_of_work = str(v(COL_NATURE_OF_WORK)).strip() if v(COL_NATURE_OF_WORK) else None
        activity = str(v(COL_ACTIVITY_TASK)).strip() if v(COL_ACTIVITY_TASK) else None
        environment = str(v(COL_ENVIRONMENT)).strip() if v(COL_ENVIRONMENT) else None
        device_id = str(v(COL_DEVICE_ID)).strip() if v(COL_DEVICE_ID) else None
        qc_dur = _safe_float(v(COL_QC_DURATION))
        exp_dur = _safe_float(v(COL_EXPECTED_DURATION))
        total_dur = _safe_float(v(COL_TOTAL_DURATION))
        payout = _safe_float(v(COL_TOTAL_PAYOUT))

        # Compute total_duration from QC + Expected if missing
        if total_dur is None and qc_dur is not None and exp_dur is not None:
            total_dur = qc_dur + exp_dur
        elif total_dur is None and qc_dur is not None:
            total_dur = qc_dur
        elif total_dur is None and exp_dur is not None:
            total_dur = exp_dur

        # Store raw data for traceability
        raw_data = {}
        for i, val in enumerate(values):
            if val is not None:
                raw_data[get_column_letter(i + 1)] = str(val)

        return ImportRow(
            row_number=row_number,
            participant_id=participant_id,
            nric_passport=nric,
            full_name=full_name,
            short_name=short_name,
            phone=phone,
            date=parsed_date.isoformat() if parsed_date else None,
            project_id=project_id,
            project_session=project_session,
            body_height_cm=body_height,
            nature_of_work=nature_of_work,
            activity_task=activity,
            environment=environment,
            device_id=device_id,
            qc_duration_minutes=qc_dur,
            expected_duration_minutes=exp_dur,
            total_duration_minutes=total_dur,
            total_payout_rm=payout,
            raw_data=raw_data,
            is_valid=True,
            warnings=[],
            errors=[],
        )


class RowValidator:
    """Validates normalized ImportRow objects."""

    @staticmethod
    def validate(row: ImportRow) -> ImportRow:
        errors = []
        warnings = []

        if not row.participant_id:
            errors.append("Missing Participant ID")
        if not row.full_name:
            errors.append("Missing Full Name")
        if not row.date:
            errors.append("Invalid or missing Date")
        if row.total_duration_minutes is not None and row.total_duration_minutes < 0:
            errors.append(f"Negative duration: {row.total_duration_minutes} min")
        if row.total_payout_rm is not None and row.total_payout_rm < 0:
            errors.append(f"Negative payout: RM {row.total_payout_rm}")

        if not row.phone:
            warnings.append("Missing Phone")
        if not row.nric_passport:
            warnings.append("Missing NRIC/Passport")
        if not row.project_session:
            warnings.append("Missing Project Session")
        if not row.device_id:
            warnings.append("Missing Device ID")

        row.errors = errors
        row.warnings = warnings
        row.is_valid = len(errors) == 0
        return row


class DuplicateDetector:
    """Detects duplicates by checking existing imported records in the database."""

    def __init__(self, db: AsyncSession):
        self.db = db
        self._existing_sessions: dict = {}
        self._existing_users: dict = {}
        self._loaded = False

    async def _load_existing(self):
        if self._loaded:
            return

        result = await self.db.execute(
            select(TaskSession).where(TaskSession.source == DataSource.IMPORTED)
        )
        for session in result.scalars().all():
            if session.import_reference:
                self._existing_sessions[session.import_reference] = session

        result2 = await self.db.execute(
            select(User).where(
                User.source == DataSource.IMPORTED,
                User.legacy_participant_id.isnot(None)
            )
        )
        for user in result2.scalars().all():
            if user.legacy_participant_id:
                self._existing_users[user.legacy_participant_id] = user

        self._loaded = True

    async def classify(self, row: ImportRow) -> tuple[str, list[str]]:
        """
        Classify row status: valid | duplicate_exact | duplicate_possible | conflict
        """
        await self._load_existing()
        warnings = []

        if not row.participant_id or not row.project_session:
            return "valid", warnings

        existing = self._existing_sessions.get(row.project_session)
        if existing:
            existing_date = existing.checked_in_at.date() if existing.checked_in_at else None
            row_date = None
            if row.date:
                try:
                    row_date = datetime.fromisoformat(row.date).date()
                except (ValueError, TypeError):
                    pass

            same_date = existing_date == row_date
            same_duration = (
                existing.duration_minutes is not None
                and row.total_duration_minutes is not None
                and abs(existing.duration_minutes - row.total_duration_minutes) < 0.5
            )

            if same_date and same_duration:
                return "duplicate_exact", warnings
            if same_date:
                return "duplicate_possible", ["Different duration/payout for same participant + date"]

            worker_result = await self.db.execute(
                select(User).where(User.id == existing.worker_id)
            )
            existing_worker = worker_result.scalar_one_or_none()
            if existing_worker and existing_worker.legacy_participant_id != row.participant_id:
                return "conflict", [f"Session {row.project_session} already belongs to participant {existing_worker.legacy_participant_id}"]

        return "valid", warnings

    async def find_existing_worker(self, participant_id: str) -> Optional[User]:
        await self._load_existing()
        return self._existing_users.get(participant_id)


class ImportService:
    """
    Orchestrates the full import pipeline with preview and confirm phases.
    """

    def __init__(self, db: AsyncSession):
        self.db = db
        self.parser = ExcelRowParser()
        self.validator = RowValidator()
        self.detector = DuplicateDetector(db)

    # ── Parsing ────────────────────────────────────────────────────────────

    def parse_workbook(
        self,
        file_content: bytes,
        filename: str,
        worksheet_name: Optional[str] = None,
    ) -> tuple[list[ImportRow], str, str]:
        """Parse Excel into ImportRow list. Returns (rows, filename, sheet_name)."""
        wb = load_workbook(io.BytesIO(file_content), data_only=True)

        try:
            if worksheet_name and worksheet_name in wb.sheetnames:
                ws = wb[worksheet_name]
            elif "Feb - June Tracker (Cleaned)" in wb.sheetnames:
                ws = wb["Feb - June Tracker (Cleaned)"]
            else:
                ws = wb.active

            actual_worksheet = ws.title
            rows: list[ImportRow] = []

            for row_idx, row_data in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    continue
                if all(cell is None for cell in row_data):
                    continue
                parsed = self.parser.parse_row(row_data, row_idx)
                rows.append(parsed)

            return rows, filename, actual_worksheet
        finally:
            wb.close()

    # ── Validation ─────────────────────────────────────────────────────────

    def validate_rows(self, rows: list[ImportRow]) -> list[ImportRow]:
        return [self.validator.validate(row) for row in rows]

    # ── Duplicate Detection ────────────────────────────────────────────────

    async def detect_duplicates(self, rows: list[ImportRow]) -> list[ImportRow]:
        for row in rows:
            status, extra_warnings = await self.detector.classify(row)
            row.warnings.extend(extra_warnings)
            row._dup_status = status  # type: ignore[attr-defined]
        return rows

    # ── Placeholder Records ────────────────────────────────────────────────

    async def _ensure_placeholder_records(self, admin_user: User) -> tuple[uuid.UUID, uuid.UUID]:
        """
        Ensure placeholder Task and Application records exist for imported sessions.
        Creates them on first import, reuses on subsequent imports.
        """
        # Check if placeholder task already exists
        result = await self.db.execute(select(Task).where(Task.id == PLACEHOLDER_TASK_ID))
        task = result.scalar_one_or_none()
        if not task:
            task = Task(
                id=PLACEHOLDER_TASK_ID,
                employer_id=admin_user.id,
                title="Historical Import (Legacy Data)",
                description="Session imported from historical Excel workbook. See nature_of_work field for details.",
                location="Various",
                category="Historical",
                pay_rate_per_minute=0.0,
                estimated_duration_minutes=0,
                status=TaskStatus.COMPLETED,
            )
            self.db.add(task)

        # Check if placeholder application already exists
        result = await self.db.execute(select(Application).where(Application.id == PLACEHOLDER_APP_ID))
        app = result.scalar_one_or_none()
        if not app:
            app = Application(
                id=PLACEHOLDER_APP_ID,
                task_id=PLACEHOLDER_TASK_ID,
                worker_id=admin_user.id,
                status="approved",
            )
            self.db.add(app)

        return PLACEHOLDER_TASK_ID, PLACEHOLDER_APP_ID

    # ── Preview ────────────────────────────────────────────────────────────

    async def preview(
        self,
        file_content: bytes,
        filename: str,
        worksheet_name: Optional[str] = None,
    ) -> ImportPreviewResponse:
        """Preview import without making DB changes."""
        rows, actual_filename, actual_worksheet = self.parse_workbook(file_content, filename, worksheet_name)
        rows = self.validate_rows(rows)
        rows = await self.detect_duplicates(rows)

        valid_rows = [r for r in rows if r.is_valid]
        invalid_rows = [r for r in rows if not r.is_valid]
        exact_dupes = [r for r in rows if getattr(r, '_dup_status', 'valid') == 'duplicate_exact']
        possible_dupes = [r for r in rows if getattr(r, '_dup_status', 'valid') == 'duplicate_possible']
        conflicts = [r for r in rows if getattr(r, '_dup_status', 'valid') == 'conflict']

        seen_pids = set()
        workers_to_create = 0
        workers_matched = 0
        worker_previews = []

        for row in valid_rows:
            if row.participant_id and row.participant_id not in seen_pids:
                seen_pids.add(row.participant_id)
                existing = await self.detector.find_existing_worker(row.participant_id)
                if existing:
                    workers_matched += 1
                    worker_previews.append(WorkerPreview(
                        participant_id=row.participant_id,
                        full_name=row.full_name or "",
                        phone=row.phone,
                        nric_passport=row.nric_passport,
                        status="existing",
                    ))
                else:
                    workers_to_create += 1
                    worker_previews.append(WorkerPreview(
                        participant_id=row.participant_id,
                        full_name=row.full_name or "",
                        phone=row.phone,
                        nric_passport=row.nric_passport,
                        status="new",
                    ))

        all_warnings = set()
        all_errors = set()
        for r in rows:
            for w in r.warnings:
                all_warnings.add(w)
            for e in r.errors:
                all_errors.add(e)

        preview_rows = [
            ImportPreviewRow(
                row_number=r.row_number,
                participant_id=r.participant_id,
                full_name=r.full_name,
                date=r.date,
                project_session=r.project_session,
                duration_minutes=r.total_duration_minutes,
                earnings=r.total_payout_rm,
                status=getattr(r, '_dup_status', 'invalid' if not r.is_valid else 'valid'),
                warnings=r.warnings,
                errors=r.errors,
            )
            for r in rows
        ]

        sessions_to_import = len(valid_rows) - len(exact_dupes)

        return ImportPreviewResponse(
            filename=actual_filename,
            worksheet_name=actual_worksheet,
            total_rows=len(rows),
            valid_rows=len(valid_rows),
            invalid_rows=len(invalid_rows),
            exact_duplicates=len(exact_dupes),
            possible_duplicates=len(possible_dupes),
            conflicts=len(conflicts),
            workers_to_create=workers_to_create,
            workers_matched=workers_matched,
            sessions_to_import=sessions_to_import,
            missing_required_fields=sum(1 for r in invalid_rows if r.errors),
            validation_warnings=sorted(all_warnings),
            validation_errors=sorted(all_errors),
            workers=worker_previews,
            rows=preview_rows,
        )

    # ── Confirm Import ────────────────────────────────────────────────────

    async def confirm(
        self,
        file_content: bytes,
        filename: str,
        admin_user: User,
        worksheet_name: Optional[str] = None,
        import_version: Optional[str] = None,
    ) -> ImportConfirmResponse:
        """
        Execute import inside a single database transaction.
        On failure: saves ImportLog in a separate session to preserve audit trail.
        """
        start_time = time.time()

        # Create ImportLog record (will be saved outside main transaction on failure)
        log_record = ImportLog(
            filename=filename,
            worksheet_name=worksheet_name,
            import_version=import_version or settings.IMPORT_VERSION,
            imported_by_id=admin_user.id,
            status="running",
            total_rows=0,
            valid_rows=0,
            duplicate_rows=0,
            workers_created=0,
            workers_matched=0,
            sessions_imported=0,
        )

        try:
            # Parse, validate, detect duplicates
            rows, actual_filename, actual_worksheet = self.parse_workbook(
                file_content, filename, worksheet_name
            )
            rows = self.validate_rows(rows)
            rows = await self.detect_duplicates(rows)

            valid_rows = [r for r in rows if r.is_valid]
            exact_dupes = [r for r in rows if getattr(r, '_dup_status', 'valid') == 'duplicate_exact']
            possible_dupes = [r for r in rows if getattr(r, '_dup_status', 'valid') == 'duplicate_possible']
            conflicts = [r for r in rows if getattr(r, '_dup_status', 'valid') == 'conflict']

            importable = [
                r for r in valid_rows
                if getattr(r, '_dup_status', 'valid') not in ('duplicate_exact', 'conflict')
            ]

            # Ensure placeholder task/application records exist
            ph_task_id, ph_app_id = await self._ensure_placeholder_records(admin_user)

            # Create/Find workers
            worker_ids_by_pid: dict[str, uuid.UUID] = {}
            workers_created = 0
            workers_matched = 0

            unique_pids = set()
            for r in importable:
                if r.participant_id:
                    unique_pids.add(r.participant_id)

            for pid in unique_pids:
                existing = await self.detector.find_existing_worker(pid)
                if existing:
                    worker_ids_by_pid[pid] = existing.id
                    workers_matched += 1
                else:
                    new_user = User(
                        email=generate_import_email(pid),
                        full_name=next(
                            (r.full_name for r in importable if r.participant_id == pid),
                            f"Imported Worker {pid}"
                        ) or f"Imported Worker {pid}",
                        source=DataSource.IMPORTED,
                        legacy_participant_id=pid,
                        is_active=False,
                        is_verified=False,
                        verification_status="imported",
                        phone=next(
                            (r.phone for r in importable if r.participant_id == pid and r.phone),
                            None
                        ),
                        nric_passport=next(
                            (r.nric_passport for r in importable if r.participant_id == pid and r.nric_passport),
                            None
                        ),
                        body_height_cm=next(
                            (r.body_height_cm for r in importable if r.participant_id == pid and r.body_height_cm),
                            None
                        ),
                    )
                    self.db.add(new_user)
                    await self.db.flush()
                    await self.db.refresh(new_user)
                    worker_ids_by_pid[pid] = new_user.id
                    workers_created += 1

            # Create TaskSession records
            sessions_imported = 0
            failed_rows_details = []

            for r in importable:
                worker_id = worker_ids_by_pid.get(r.participant_id) if r.participant_id else None
                if not worker_id:
                    failed_rows_details.append({
                        "row_number": r.row_number,
                        "error": "No matching worker found",
                    })
                    continue

                parsed_date = None
                if r.date:
                    try:
                        parsed_date = datetime.fromisoformat(r.date).replace(
                            hour=8, minute=0, tzinfo=timezone.utc
                        )
                    except (ValueError, TypeError):
                        parsed_date = None

                # Use the task title/description from Excel's nature_of_work and activity fields
                # stored in proof_notes so existing queries can show them
                proof_notes_parts = []
                if r.nature_of_work:
                    proof_notes_parts.append(f"Work: {r.nature_of_work}")
                if r.activity_task:
                    proof_notes_parts.append(f"Activity: {r.activity_task}")
                if r.project_session:
                    proof_notes_parts.append(f"Ref: {r.project_session}")

                session = TaskSession(
                    worker_id=worker_id,
                    task_id=ph_task_id,
                    application_id=ph_app_id,
                    checked_in_at=parsed_date or datetime(2026, 1, 1, 8, 0, tzinfo=timezone.utc),
                    checked_out_at=parsed_date,
                    earnings=r.total_payout_rm,
                    status=SessionStatus.SETTLED,
                    source=DataSource.IMPORTED,
                    import_reference=r.project_session,
                    duration_minutes=r.total_duration_minutes,
                    nature_of_work=r.nature_of_work,
                    work_environment=r.environment,
                    legacy_device_id=r.device_id,
                    raw_import_data=r.raw_data,
                    proof_notes=" | ".join(proof_notes_parts) if proof_notes_parts else None,
                )
                self.db.add(session)
                sessions_imported += 1

            # Update ImportLog
            log_record.total_rows = len(rows)
            log_record.valid_rows = len(valid_rows)
            log_record.duplicate_rows = len(exact_dupes)
            log_record.workers_created = workers_created
            log_record.workers_matched = workers_matched
            log_record.sessions_imported = sessions_imported
            log_record.failed_rows_details = failed_rows_details if failed_rows_details else None
            log_record.status = "completed"
            log_record.completed_at = datetime.now(timezone.utc)

            self.db.add(log_record)
            await self.db.flush()
            await self.db.refresh(log_record)

            elapsed = round(time.time() - start_time, 2)

            return ImportConfirmResponse(
                status="completed",
                filename=actual_filename,
                worksheet_name=actual_worksheet,
                total_rows=len(rows),
                valid_rows=len(valid_rows),
                workers_created=workers_created,
                workers_matched=workers_matched,
                sessions_imported=sessions_imported,
                exact_duplicates=len(exact_dupes),
                possible_duplicates=len(possible_dupes),
                conflicts=len(conflicts),
                failed_rows=len(failed_rows_details),
                execution_time_seconds=elapsed,
                import_log_id=str(log_record.id),
            )

        except Exception as e:
            logger.exception("Import failed: %s", e)
            elapsed = round(time.time() - start_time, 2)

            # Save failed ImportLog in a SEPARATE session to survive rollback
            log_record.status = "failed"
            log_record.error_log = {"error": str(e), "detail": f"Import failed after {elapsed}s"}
            log_record.completed_at = datetime.now(timezone.utc)

            try:
                from app.database import AsyncSessionLocal
                async with AsyncSessionLocal() as fail_db:
                    fail_db.add(log_record)
                    await fail_db.commit()
                logger.info("Failed import logged to import_logs (id=%s)", log_record.id)
            except Exception as inner_e:
                logger.error("Failed to save failed ImportLog: %s", inner_e)

            raise